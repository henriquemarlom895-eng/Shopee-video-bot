"""
Cria a planilha modelo planilha.xlsx com cabeçalho e exemplos.
Execute uma vez: python criar_planilha.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Promoções"

# Cabeçalho
headers = ["Link Shopee (afiliado)", "Texto extra (opcional)", "Status"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="E53935")
    cell.alignment = Alignment(horizontal="center")

# Larguras
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 25

# Linha de exemplo
ws["A2"] = "https://shope.ee/SEU_LINK_AQUI"
ws["B2"] = ""
ws["C2"] = ""

wb.save("planilha.xlsx")
print("✅ planilha.xlsx criada com sucesso!")
