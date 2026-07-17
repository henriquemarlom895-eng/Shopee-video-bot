"""
Bot de promoções — lê planilha Excel e publica no Facebook, Instagram e Threads.

Uso:
  python main.py

A planilha deve ter:
  Coluna A: Link da Shopee (afiliado ou normal)
  Coluna B: Texto extra opcional (deixe em branco para usar o modelo padrão)
  Coluna C: Status (preenchido automaticamente pelo script)
"""
import os
import sys
import time

import openpyxl
from dotenv import load_dotenv

from shopee import get_product_info
from meta import MetaPoster

# ── Configuração ──────────────────────────────────────────────────────────────
load_dotenv()

FB_PAGE_ID     = os.environ.get("FB_PAGE_ID", "")
FB_PAGE_TOKEN  = os.environ.get("FB_PAGE_TOKEN", "")
IG_USER_ID     = os.environ.get("IG_USER_ID", "")
THREADS_USER_ID = os.environ.get("THREADS_USER_ID", "")
PLANILHA       = os.environ.get("PLANILHA", "planilha.xlsx")

TEMPLATE = os.environ.get(
    "TEMPLATE_TEXTO",
    "🔥 {nome}\n\n💰 Por apenas R$ {preco}\n\n🛒 Compre agora: {link}\n\n#shopee #promoção #oferta #desconto"
)

# ── Validação ─────────────────────────────────────────────────────────────────
erros = []
if not FB_PAGE_ID or not FB_PAGE_TOKEN:
    erros.append("FB_PAGE_ID e FB_PAGE_TOKEN são obrigatórios no .env")
if not IG_USER_ID:
    erros.append("IG_USER_ID é obrigatório no .env")

if erros:
    print("❌ Configuração incompleta:\n" + "\n".join(f"  • {e}" for e in erros))
    print("\nCopie o arquivo .env.example para .env e preencha os valores.")
    sys.exit(1)

if not os.path.exists(PLANILHA):
    print(f"❌ Planilha '{PLANILHA}' não encontrada.")
    print("Crie uma planilha com os links na coluna A.")
    sys.exit(1)

poster = MetaPoster(FB_PAGE_ID, FB_PAGE_TOKEN, IG_USER_ID, THREADS_USER_ID)


# ── Helpers ───────────────────────────────────────────────────────────────────

def formatar_preco(preco: float) -> str:
    return f"{preco:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def gerar_texto(nome: str, preco: float, link: str, texto_extra: str = "") -> str:
    base = TEMPLATE.format(
        nome=nome[:120],
        preco=formatar_preco(preco),
        link=link,
    )
    if texto_extra:
        base = texto_extra.strip() + "\n\n" + base
    return base


# ── Processamento ─────────────────────────────────────────────────────────────

def processar_planilha():
    wb = openpyxl.load_workbook(PLANILHA)
    ws = wb.active

    # Cabeçalho (se ainda não existir)
    if ws["A1"].value in (None, ""):
        ws["A1"] = "Link Shopee"
        ws["B1"] = "Texto extra (opcional)"
        ws["C1"] = "Status"
        wb.save(PLANILHA)

    publicados = 0
    erros_total = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        link_cell   = row[0]
        extra_cell  = row[1] if len(row) > 1 else None
        status_cell = row[2] if len(row) > 2 else ws.cell(row=idx, column=3)

        link = str(link_cell.value or "").strip()
        if not link or not link.startswith("http"):
            continue

        status = str(status_cell.value or "").strip()
        if status == "Publicado":
            print(f"⏭️  Linha {idx}: já publicado, pulando.")
            continue

        print(f"\n{'─'*50}")
        print(f"📦 Linha {idx}: {link}")

        # Busca produto
        info = get_product_info(link)
        if not info:
            status_cell.value = "Erro — produto não encontrado"
            wb.save(PLANILHA)
            erros_total += 1
            continue

        nome       = info["name"]
        preco      = info["price"]
        image_url  = info["image_url"]
        texto_extra = str(extra_cell.value or "").strip() if extra_cell else ""

        print(f"  📝 {nome} — R$ {formatar_preco(preco)}")
        print(f"  🖼️  {image_url}")

        texto = gerar_texto(nome, preco, link, texto_extra)
        resultados = []

        # ── Facebook ──
        print("  📘 Postando no Facebook...")
        r = poster.post_facebook(texto, image_url)
        ok_fb = "id" in r
        print(f"     {'✅' if ok_fb else '❌'} {r}")
        resultados.append(ok_fb)
        time.sleep(3)

        # ── Instagram Feed ──
        print("  📸 Postando no Instagram (feed)...")
        r = poster.post_instagram_feed(texto, image_url)
        ok_ig = "id" in r
        print(f"     {'✅' if ok_ig else '❌'} {r}")
        resultados.append(ok_ig)
        time.sleep(3)

        # ── Instagram Story ──
        if image_url:
            print("  📖 Postando nos Stories do Instagram...")
            r = poster.post_instagram_story(image_url)
            ok_story = "id" in r
            print(f"     {'✅' if ok_story else '❌'} {r}")
            resultados.append(ok_story)
            time.sleep(3)

        # ── Threads ──
        if THREADS_USER_ID:
            print("  🧵 Postando no Threads (aguarda 32s)...")
            r = poster.post_threads(texto, image_url)
            ok_th = "id" in r
            print(f"     {'✅' if ok_th else '❌'} {r}")
            resultados.append(ok_th)

        # Atualiza status
        if all(resultados):
            status_cell.value = "Publicado"
            publicados += 1
        elif any(resultados):
            status_cell.value = "Publicado parcialmente"
            publicados += 1
        else:
            status_cell.value = "Erro — falha ao publicar"
            erros_total += 1

        wb.save(PLANILHA)
        time.sleep(10)  # Pausa entre produtos para evitar bloqueios

    print(f"\n{'═'*50}")
    print(f"✅ Concluído! {publicados} publicados, {erros_total} erros.")


if __name__ == "__main__":
    print("🚀 Bot de promoções iniciando...")
    processar_planilha()
